from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

# from flask_bcrypt import Bcrypt
from flask_jwt_extended import (
    JWTManager,
    create_access_token,
    jwt_required,
    get_jwt_identity,
)
from werkzeug.utils import secure_filename

import pandas as pd
import os
import uuid
import re
import secrets
import string
from datetime import datetime, timedelta
import mysql.connector
from mysql.connector import Error
import hashlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO, StringIO

app = Flask(__name__)
CORS(app)
# bcrypt = Bcrypt(app)
jwt = JWTManager(app)

# Configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
PROCESSED_FOLDER = os.path.join(BASE_DIR, "processed")
ALLOWED_EXTENSIONS = {"xlsx", "xls"}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB
SUPER_SECRET_KEY = "1000super-secret-key"


os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["PROCESSED_FOLDER"] = PROCESSED_FOLDER
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE
app.config["JWT_SECRET_KEY"] = SUPER_SECRET_KEY

# Configuration base de données MySQL
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "excel_processor_db",
}


# ==================== GESTION BASE DE DONNÉES ====================


def get_db_connection():
    """Établit une connexion à la base de données MySQL"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        print(f"Erreur de connexion à MySQL: {e}")
        return None


def init_database():
    """Initialise la base de données et crée les tables nécessaires"""
    try:
        connection = mysql.connector.connect(
            host=DB_CONFIG["host"],
            user=DB_CONFIG["user"],
            password=DB_CONFIG["password"],
        )
        cursor = connection.cursor()

        # Créer la base de données si elle n'existe pas
        cursor.execute(f"CREATE DATABASE IF NOT EXISTS {DB_CONFIG['database']}")
        cursor.execute(f"USE {DB_CONFIG['database']}")

        # Table pour l'historique des fichiers
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS file_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                file_id VARCHAR(100) UNIQUE NOT NULL,
                original_filename VARCHAR(255) NOT NULL,
                processed_filename VARCHAR(255),
                upload_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                status ENUM('pending', 'processing', 'success', 'error') DEFAULT 'pending',
                record_count INT DEFAULT 0,
                error_message TEXT,
                user_id VARCHAR(100),
                INDEX idx_file_id (file_id),
                INDEX idx_status (status),
                INDEX idx_upload_date (upload_date)
            )
        """
        )

        # Table pour les données traitées
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS processed_records (
                id INT AUTO_INCREMENT PRIMARY KEY,
                file_id VARCHAR(100) NOT NULL,
                nom VARCHAR(100) NOT NULL,
                prenom VARCHAR(100) NOT NULL,
                numero VARCHAR(50) NOT NULL,
                email VARCHAR(255) NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (file_id) REFERENCES file_history(file_id) ON DELETE CASCADE,
                INDEX idx_file_id (file_id),
                INDEX idx_email (email)
            )
        """
        )

        # Table pour les utilisateurs (optionnel - pour l'authentification)
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                email VARCHAR(255) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                role VARCHAR(255) NOT NULL DEFAULT 'operator',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_login DATETIME,
                is_active BOOLEAN DEFAULT TRUE,
                INDEX idx_username (username),
                INDEX idx_email (email)
            )
        """
        )

        connection.commit()
        cursor.close()
        connection.close()
        print("Base de données initialisée avec succès!")

    except Error as e:
        print(f"Erreur lors de l'initialisation de la base de données: {e}")


# ==================== UTILITAIRES ====================


def allowed_file(filename):
    """Vérifie si l'extension du fichier est autorisée"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_email(prenom, nom):
    """Génère un email au format prenom.nom@entreprise.com"""
    # Normalisation : suppression des accents et caractères spéciaux
    prenom_clean = re.sub(
        r"[^a-zA-Z]",
        "",
        prenom.lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
        .replace("î", "i")
        .replace("ç", "c"),
    )
    nom_clean = re.sub(
        r"[^a-zA-Z]",
        "",
        nom.lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
        .replace("ô", "o")
        .replace("î", "i")
        .replace("ç", "c"),
    )

    return f"{prenom_clean}.{nom_clean}@entreprise.com"


def generate_strong_password(length=12):
    """Génère un mot de passe fort avec majuscules, minuscules, chiffres et symboles"""
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"

    # S'assurer qu'il y a au moins un de chaque type
    password = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%^&*"),
    ]

    # Compléter avec des caractères aléatoires
    password += [secrets.choice(alphabet) for _ in range(length - 4)]

    # Mélanger
    secrets.SystemRandom().shuffle(password)

    return "".join(password)


def hash_password(password):
    """Hash un mot de passe avec SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()


def validate_phone_number(numero):
    """Valide le format d'un numéro de téléphone"""
    numero_clean = re.sub(r"\s+", "", str(numero))
    pattern = r"^\+?[0-9]{8,15}$"
    return re.match(pattern, numero_clean) is not None


def validate_excel_data(df):
    """Valide les données du fichier Excel"""
    errors = []
    required_columns = ["nom", "prenom", "numero"]

    # Vérifier les colonnes requises
    df_columns = [col.lower() for col in df.columns]
    missing_columns = [col for col in required_columns if col not in df_columns]

    if missing_columns:
        errors.append(f"Colonnes manquantes: {', '.join(missing_columns)}")
        return False, errors

    # Normaliser les noms de colonnes
    df.columns = [col.lower() for col in df.columns]

    # Vérifier les doublons et les valeurs manquantes
    numeros_seen = set()
    emails_generated = set()

    for index, row in df.iterrows():
        row_num = index + 2  # +2 car Excel commence à 1 et la ligne 1 est l'en-tête

        # Vérifier les valeurs manquantes
        if pd.isna(row["nom"]) or str(row["nom"]).strip() == "":
            errors.append(f"Ligne {row_num}: Nom manquant")

        if pd.isna(row["prenom"]) or str(row["prenom"]).strip() == "":
            errors.append(f"Ligne {row_num}: Prénom manquant")

        if pd.isna(row["numero"]) or str(row["numero"]).strip() == "":
            errors.append(f"Ligne {row_num}: Numéro manquant")
        else:
            numero = str(row["numero"]).strip()

            # Vérifier les doublons
            if numero in numeros_seen:
                errors.append(f"Ligne {row_num}: Numéro en double ({numero})")
            numeros_seen.add(numero)

            # Vérifier le format
            if not validate_phone_number(numero):
                errors.append(f"Ligne {row_num}: Format de numéro invalide ({numero})")

        # Vérifier les emails potentiellement en double
        if not pd.isna(row["nom"]) and not pd.isna(row["prenom"]):
            email = generate_email(str(row["prenom"]), str(row["nom"]))
            if email in emails_generated:
                errors.append(f"Ligne {row_num}: Email en double potentiel ({email})")
            emails_generated.add(email)

    return len(errors) == 0, errors


# ==================== ROUTES API ====================


# ---------------------------
# 🔹 Inscription
# ---------------------------
@app.route("/api/auth/register", methods=["POST"])
def register():
    data = request.get_json()
    username = data.get("username")
    email = data.get("email")
    password = data.get("password")
    role = data.get("role", "operator")

    if not username or not email or not password:
        return jsonify({"message": "Tous les champs sont obligatoires."}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Vérifier si l'utilisateur existe déjà
        cursor.execute(
            "SELECT * FROM users WHERE email = %s OR username = %s", (email, username)
        )
        if cursor.fetchone():
            return jsonify({"message": "Cet utilisateur existe déjà."}), 400

        # Hasher le mot de passe
        hashed_password = (
            password  # bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
        )

        # Insérer dans la table
        cursor.execute(
            """
            INSERT INTO users (username, email, password_hash,role)
            VALUES (%s, %s, %s, %s)
        """,
            (username, email, hashed_password, role),
        )
        conn.commit()

        return (
            jsonify(
                {
                    "message": "Inscription réussie.",
                    "validation": "true",
                }
            ),
            201,
        )

    except Exception as e:
        print("Erreur:", e)
        return jsonify({"message": "Erreur lors de l’inscription."}), 500

    finally:
        cursor.close()
        conn.close()


# ---------------------------
# 🔹 Connexion
# ---------------------------
@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json()
    email = data.get("email")
    password = data.get("password")

    if not email or not password:
        return jsonify({"message": "Email et mot de passe requis."}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()

        if not user:
            return jsonify({"message": "Utilisateur non trouvé."}), 404

        # Vérifier le mot de passe
        # if not bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
        #     return jsonify({'message': 'Mot de passe incorrect.'}), 401
        if password != user["password_hash"]:
            return jsonify({"message": "Mot de passe incorrect."}), 401

        # Mettre à jour la date de dernière connexion
        cursor.execute(
            "UPDATE users SET last_login = NOW() WHERE id = %s", (user["id"],)
        )
        conn.commit()

        # Créer un token JWT
        access_token = create_access_token(
            identity=email, expires_delta=timedelta(hours=24)
        )

        return (
            jsonify(
                {
                    "message": "Connexion réussie.",
                    "token": access_token,
                    "user": {
                        "id": user["id"],
                        "username": user["username"],
                        "email": user["email"],
                        "role": user["role"],
                    },
                }
            ),
            200,
        )

    except Exception as e:
        print("Erreur:", e)
        return jsonify({"message": "Erreur lors de la connexion."}), 500

    finally:
        cursor.close()
        conn.close()


# ---------------------------
# 🔹 change mot de passe
# ---------------------------
@app.route("/api/auth/change_password", methods=["POST"])
@jwt_required()
def change_password():
    data = request.get_json()
    old_password = data.get("old_password")
    new_password = data.get("new_password")

    if not old_password or not new_password:
        return jsonify({"message": "Ancien et nouveau mot de passe requis."}), 400

    try:
        current_user_email = get_jwt_identity()
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM users WHERE email = %s", (current_user_email,))
        user = cursor.fetchone()

        if not user:
            return jsonify({"message": "Utilisateur non trouvé."}), 404

        # Vérifier l'ancien mot de passe
        if old_password != user["password_hash"]:
            return jsonify({"message": "Ancien mot de passe incorrect."}), 401

        # Mettre à jour le mot de passe
        cursor.execute(
            "UPDATE users SET password_hash = %s WHERE id = %s",
            (new_password, user["id"]),
        )
        conn.commit()

        return jsonify({"message": "Mot de passe mis à jour avec succès."}), 200

    except Exception as e:
        print("Erreur:", e)
        return jsonify({"message": "Erreur lors du changement de mot de passe."}), 500

    finally:
        cursor.close()
        conn.close()


@app.route("/api/health", methods=["GET"])
def health_check():
    """Vérification de l'état du serveur"""
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat()})


@app.route("/api/Tableau_bord_admin", methods=["GET"])
@jwt_required()
def get_all_users():
    # L'identité JWT est juste l'email
    current_user_email = get_jwt_identity()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Récupérer l'utilisateur connecté
    cursor.execute("SELECT id, role FROM users WHERE email = %s", (current_user_email,))
    current_user = cursor.fetchone()

    if not current_user:
        cursor.close()
        conn.close()
        return jsonify({"message": "Utilisateur non trouvé"}), 404

    # Vérifier si c’est bien un admin
    if current_user["role"] != "admin":
        cursor.close()
        conn.close()
        return jsonify({"message": "Accès refusé : administrateur requis"}), 403

    # Récupérer tous les autres utilisateurs SAUF l’admin connecté
    cursor.execute(
        "SELECT id, username, email, role, last_login FROM users WHERE deleted = FALSE AND id != %s",
        (current_user["id"],),
    )
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify(users)


@app.route("/api/users/<int:user_id>/delete", methods=["PUT"])
@jwt_required()
def delete_user(user_id):
    current_user = get_jwt_identity()
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Vérifier si l'utilisateur est admin
    cursor.execute("SELECT role FROM users WHERE email = %s", (current_user,))
    admin = cursor.fetchone()

    if not admin or admin["role"] != "admin":
        cursor.close()
        conn.close()
        return jsonify({"message": "Accès refusé : administrateur requis"}), 403

    try:
        # Marquer l'utilisateur comme supprimé
        cursor.execute("UPDATE users SET deleted = TRUE WHERE id = %s", (user_id,))
        conn.commit()

        cursor.close()
        conn.close()
        return jsonify({"message": "Utilisateur supprimé avec succès"}), 200

    except Exception as e:
        print("Erreur suppression utilisateur:", e)
        cursor.close()
        conn.close()
        return jsonify({"message": "Erreur lors de la suppression"}), 500



# Modification de la route /api/upload pour traiter toutes les feuilles


@app.route("/api/upload", methods=["POST"])
def upload_file():
    """Upload et traitement d'un fichier Excel (toutes les feuilles)"""

    if "file" not in request.files:
        return jsonify({"error": "Aucun fichier fourni"}), 400

    file = request.files["file"]
    custom_filename = request.form.get("custom_filename", None)

    if file.filename == "":
        return jsonify({"error": "Nom de fichier vide"}), 400

    if not allowed_file(file.filename):
        return (
            jsonify({"error": "Type de fichier non autorisé. Utilisez .xlsx ou .xls"}),
            400,
        )

    try:
        file_id = str(uuid.uuid4())
        original_filename = secure_filename(file.filename)
        upload_path = os.path.join(
            app.config["UPLOAD_FOLDER"], f"{file_id}_{original_filename}"
        )
        file.save(upload_path)

        # Enregistrer dans la base de données
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            cursor.execute(
                """
                INSERT INTO file_history (file_id, original_filename, status)
                VALUES (%s, %s, 'processing')
            """,
                (file_id, original_filename),
            )
            connection.commit()
            cursor.close()
            connection.close()

        # Lire TOUTES les feuilles du fichier Excel
        excel_file = pd.ExcelFile(upload_path)
        all_sheets_data = []
        validation_errors = []

        # Traiter chaque feuille
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(upload_path, sheet_name=sheet_name)

            # Valider les données de cette feuille
            is_valid, sheet_errors = validate_excel_data(df)

            if not is_valid:
                # Ajouter le nom de la feuille aux erreurs
                prefixed_errors = [
                    f"[Feuille '{sheet_name}'] {error}" for error in sheet_errors
                ]
                validation_errors.extend(prefixed_errors)
                continue

            # Traiter les données de cette feuille
            for _, row in df.iterrows():
                email = generate_email(str(row["prenom"]), str(row["nom"]))
                password = generate_strong_password()
                password_hash = hash_password(password)

                processed_record = {
                    "sheet_name": sheet_name,
                    "nom": str(row["nom"]).strip(),
                    "prenom": str(row["prenom"]).strip(),
                    "numero": str(row["numero"]).strip(),
                    "email": email,
                    "mot_de_passe": password,
                }

                all_sheets_data.append(processed_record)

                # Sauvegarder dans la base de données
                connection = get_db_connection()
                if connection:
                    cursor = connection.cursor()
                    cursor.execute(
                        """
                        INSERT INTO processed_records 
                        (file_id, nom, prenom, numero, email, password_hash)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                        (
                            file_id,
                            processed_record["nom"],
                            processed_record["prenom"],
                            processed_record["numero"],
                            processed_record["email"],
                            password_hash,
                        ),
                    )
                    connection.commit()
                    cursor.close()
                    connection.close()

        # Si toutes les feuilles ont échoué
        if validation_errors and len(all_sheets_data) == 0:
            connection = get_db_connection()
            if connection:
                cursor = connection.cursor()
                cursor.execute(
                    """
                    UPDATE file_history 
                    SET status = 'error', error_message = %s
                    WHERE file_id = %s
                """,
                    ("; ".join(validation_errors), file_id),
                )
                connection.commit()
                cursor.close()
                connection.close()

            return (
                jsonify({"error": "Validation échouée", "details": validation_errors}),
                400,
            )

        # Créer le fichier Excel traité avec TOUTES les feuilles
        if custom_filename and custom_filename.strip():
            base_filename = secure_filename(custom_filename.strip())
            if not base_filename.lower().endswith((".xlsx", ".xls")):
                base_filename += ".xlsx"
        else:
            base_filename = original_filename

        processed_filename = base_filename
        processed_path = os.path.join(
            app.config["PROCESSED_FOLDER"], processed_filename
        )

        # Créer un workbook avec plusieurs feuilles
        wb = Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut

        # Grouper les données par feuille
        sheets_dict = {}
        for record in all_sheets_data:
            sheet_name = record["sheet_name"]
            if sheet_name not in sheets_dict:
                sheets_dict[sheet_name] = []
            sheets_dict[sheet_name].append(record)

        # Créer une feuille pour chaque feuille source
        for sheet_name, records in sheets_dict.items():
            ws = wb.create_sheet(title=sheet_name)

            # En-têtes avec style
            headers = ["Nom", "Prénom", "Numéro", "Email", "Mot de passe"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Données
            for row_num, record in enumerate(records, 2):
                ws.cell(row=row_num, column=1, value=record["nom"])
                ws.cell(row=row_num, column=2, value=record["prenom"])
                ws.cell(row=row_num, column=3, value=record["numero"])
                ws.cell(row=row_num, column=4, value=record["email"])
                ws.cell(row=row_num, column=5, value=record["mot_de_passe"])

            # Ajuster la largeur des colonnes
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        wb.save(processed_path)

        # Mettre à jour le statut
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            error_msg = (
                None if not validation_errors else "; ".join(validation_errors[:5])
            )  # Limiter les erreurs
            cursor.execute(
                """
                UPDATE file_history 
                SET status = 'success', processed_filename = %s, record_count = %s, error_message = %s
                WHERE file_id = %s
            """,
                (processed_filename, len(all_sheets_data), error_msg, file_id),
            )
            connection.commit()
            cursor.close()
            connection.close()

        return (
            jsonify(
                {
                    "message": "Fichier traité avec succès",
                    "file_id": file_id,
                    "record_count": len(all_sheets_data),
                    "sheets_processed": len(sheets_dict),
                    "sheet_names": list(sheets_dict.keys()),
                    "data": all_sheets_data[:10],  # Aperçu limité
                    "warnings": validation_errors if validation_errors else None,
                }
            ),
            200,
        )

    except Exception as e:
        return (
            jsonify(
                {"error": "Erreur lors du traitement du fichier", "details": str(e)}
            ),
            500,
        )


@app.route("/api/download/<file_id>/", methods=["GET"])
def download_processed_file(file_id):
    """Télécharger un fichier traité"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({"error": "Connexion à la base de données échouée"}), 500

        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT processed_filename 
            FROM file_history 
            WHERE file_id = %s AND status = 'success'
        """,
            (file_id,),
        )
        result = cursor.fetchone()
        cursor.close()
        connection.close()

        if not result or not result["processed_filename"]:
            return jsonify({"error": "Fichier non encore traité ou introuvable"}), 404

        processed_path = os.path.join(
            app.config["PROCESSED_FOLDER"], result["processed_filename"]
        )

        if not os.path.exists(processed_path):
            return (
                jsonify(
                    {
                        "error": f"Le fichier {processed_path} est introuvable sur le serveur"
                    }
                ),
                404,
            )

        return send_file(
            processed_path,
            as_attachment=True,
            download_name=result["processed_filename"],
        )

    except Exception as e:
        import traceback

        print("Erreur serveur :", traceback.format_exc())
        return (
            jsonify({"error": "Erreur lors du téléchargement", "details": str(e)}),
            500,
        )


# Route pour obtenir la liste des feuilles d'un fichier
@app.route("/api/file/<file_id>/sheets", methods=["GET"])
def get_file_sheets(file_id):
    """Récupère la liste des feuilles d'un fichier Excel"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({"error": "Connexion à la base de données échouée"}), 500

        cursor = connection.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT processed_filename 
            FROM file_history 
            WHERE file_id = %s AND status = 'success'
        """,
            (file_id,),
        )
        result = cursor.fetchone()
        cursor.close()
        connection.close()

        if not result or not result["processed_filename"]:
            return jsonify({"error": "Fichier non trouvé"}), 404

        source_path = os.path.join(
            app.config["PROCESSED_FOLDER"], result["processed_filename"]
        )

        if not os.path.exists(source_path):
            return jsonify({"error": "Fichier introuvable sur le serveur"}), 404

        # Lire les noms des feuilles
        wb = load_workbook(source_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        return jsonify({"sheets": sheet_names, "count": len(sheet_names)}), 200

    except Exception as e:
        import traceback

        print("Erreur serveur :", traceback.format_exc())
        return (
            jsonify(
                {
                    "error": "Erreur lors de la récupération des feuilles",
                    "details": str(e),
                }
            ),
            500,
        )

@app.route("/api/download/<file_id>/<format>", methods=["GET"])
def download_processed_file_format(file_id, format):
    """Télécharger un fichier traité dans le format spécifié"""

    # Récupérer les paramètres pour CSV
    csv_sheet = request.args.get("sheet", None)  # Nom de la feuille spécifique
    csv_combine = (
        request.args.get("combine", "false").lower() == "true"
    )  # Combiner toutes les feuilles

    # Formats supportés
    SUPPORTED_FORMATS = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
        "xlsb": "application/vnd.ms-excel.sheet.binary.macroEnabled.12",
        "xls": "application/vnd.ms-excel",
        "csv": "text/csv",
        "ods": "application/vnd.oasis.opendocument.spreadsheet",
    }

    if format not in SUPPORTED_FORMATS:
        return jsonify({"error": f"Format '{format}' non supporté"}), 400

    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({"error": "Connexion à la base de données échouée"}), 500

        cursor = connection.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT processed_filename 
            FROM file_history 
            WHERE file_id = %s AND status = 'success'
        """,
            (file_id,),
        )
        result = cursor.fetchone()

        if not result or not result["processed_filename"]:
            cursor.close()
            connection.close()
            return jsonify({"error": "Fichier non trouvé"}), 404

        source_path = os.path.join(
            app.config["PROCESSED_FOLDER"], result["processed_filename"]
        )

        if not os.path.exists(source_path):
            cursor.close()
            connection.close()
            return jsonify({"error": "Fichier introuvable sur le serveur"}), 404

        wb = load_workbook(source_path)

        base_name = result["processed_filename"].rsplit(".", 1)[0]
        output_filename = f"{base_name}.{format}"

        # Conversion selon le format demandé
        if format == "xlsx":
            cursor.close()
            connection.close()
            return send_file(
                source_path,
                as_attachment=True,
                download_name=output_filename,
                mimetype=SUPPORTED_FORMATS[format],
            )

        elif format == "xlsm":
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            cursor.close()
            connection.close()
            return send_file(
                output,
                as_attachment=True,
                download_name=output_filename,
                mimetype=SUPPORTED_FORMATS[format],
            )

        elif format == "xlsb":
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            cursor.close()
            connection.close()
            return send_file(
                output,
                as_attachment=True,
                download_name=output_filename,
                mimetype=SUPPORTED_FORMATS["xlsx"],
            )

        elif format == "xls":
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            cursor.close()
            connection.close()
            return send_file(
                output,
                as_attachment=True,
                download_name=output_filename,
                mimetype=SUPPORTED_FORMATS[format],
            )

        elif format == "csv":
            cursor.close()
            connection.close()

            output = StringIO()

            if csv_combine:
                # Combiner toutes les feuilles en un seul CSV
                all_data = []
                for sheet_name in wb.sheetnames:
                    df = pd.read_excel(source_path, sheet_name=sheet_name)
                    # Ajouter une colonne pour identifier la feuille
                    df.insert(0, "Feuille", sheet_name)
                    all_data.append(df)

                # Concaténer tous les DataFrames
                combined_df = pd.concat(all_data, ignore_index=True)
                combined_df.to_csv(output, index=False, encoding="utf-8-sig", sep=";")

            elif csv_sheet:
                # Télécharger une feuille spécifique
                if csv_sheet not in wb.sheetnames:
                    return (
                        jsonify({"error": f"La feuille '{csv_sheet}' n'existe pas"}),
                        404,
                    )

                df = pd.read_excel(source_path, sheet_name=csv_sheet)
                df.to_csv(output, index=False, encoding="utf-8-sig", sep=";")
                output_filename = f"{base_name}_{csv_sheet}.csv"

            else:
                # Par défaut: première feuille seulement
                first_sheet = wb.sheetnames[0]
                df = pd.read_excel(source_path, sheet_name=first_sheet)
                df.to_csv(output, index=False, encoding="utf-8-sig", sep=";")

            csv_bytes = BytesIO(output.getvalue().encode("utf-8-sig"))
            csv_bytes.seek(0)

            return send_file(
                csv_bytes,
                as_attachment=True,
                download_name=output_filename,
                mimetype=SUPPORTED_FORMATS[format],
            )

        elif format == "ods":
            output = BytesIO()

            try:
                with pd.ExcelWriter(output, engine="odf") as writer:
                    for sheet_name in wb.sheetnames:
                        df = pd.read_excel(source_path, sheet_name=sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                output.seek(0)
                cursor.close()
                connection.close()
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype=SUPPORTED_FORMATS[format],
                )
            except Exception as e:
                print(f"Erreur ODF: {e}, tentative avec XLSX...")
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                cursor.close()
                connection.close()
                return send_file(
                    output,
                    as_attachment=True,
                    download_name=output_filename.replace(".ods", ".xlsx"),
                    mimetype=SUPPORTED_FORMATS["xlsx"],
                )

    except Exception as e:
        import traceback

        print("Erreur serveur :", traceback.format_exc())
        if "cursor" in locals():
            cursor.close()
        if "connection" in locals():
            connection.close()
        return (
            jsonify(
                {
                    "error": "Erreur lors de la conversion/téléchargement",
                    "details": str(e),
                }
            ),
            500,
        )


# Route pour obtenir les formats disponibles
@app.route("/api/export-formats", methods=["GET"])
def get_export_formats():
    """Retourne la liste des formats d'export disponibles"""
    formats = [
        {
            "value": "xlsx",
            "label": "Excel Workbook (.xlsx)",
            "description": "Format moderne recommandé",
            "icon": "📊",
        },
        {
            "value": "xlsm",
            "label": "Excel Macro-Enabled (.xlsm)",
            "description": "Excel avec support des macros",
            "icon": "📈",
        },
        {
            "value": "xlsb",
            "label": "Excel Binary (.xlsb)",
            "description": "Format binaire (converti depuis xlsx)",
            "icon": "💾",
        },
        {
            "value": "xls",
            "label": "Excel 97-2003 (.xls)",
            "description": "Format ancien (converti depuis xlsx)",
            "icon": "📋",
        },
        {
            "value": "csv",
            "label": "CSV (.csv)",
            "description": "Fichier texte séparé par point-virgule",
            "icon": "📄",
        },
        {
            "value": "ods",
            "label": "OpenDocument (.ods)",
            "description": "Format LibreOffice/OpenOffice",
            "icon": "📑",
        },
    ]
    return jsonify({"formats": formats}), 200


@app.route("/api/history", methods=["GET"])
def get_history():
    """Récupérer l'historique des fichiers traités"""
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT file_id, original_filename, processed_filename, 
                       upload_date, status, record_count, error_message
                FROM file_history
                ORDER BY upload_date DESC
                LIMIT 100
            """
            )
            history = cursor.fetchall()
            cursor.close()
            connection.close()

            # Convertir les dates en format ISO
            for record in history:
                if record["upload_date"]:
                    record["upload_date"] = record["upload_date"].isoformat()

            return jsonify({"history": history}), 200

        return jsonify({"error": "Erreur de connexion à la base de données"}), 500

    except Exception as e:
        return (
            jsonify(
                {
                    "error": "Erreur lors de la récupération de l'historique",
                    "details": str(e),
                }
            ),
            500,
        )


@app.route("/api/file/<file_id>", methods=["GET"])
def get_file_details(file_id):
    """Récupérer les détails d'un fichier spécifique"""
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)

            # Récupérer les informations du fichier
            cursor.execute(
                """
                SELECT * FROM file_history WHERE file_id = %s
            """,
                (file_id,),
            )
            file_info = cursor.fetchone()

            if not file_info:
                return jsonify({"error": "Fichier non trouvé"}), 404

            # Récupérer les enregistrements traités
            cursor.execute(
                """
                SELECT nom, prenom, numero, email, created_at
                FROM processed_records
                WHERE file_id = %s
            """,
                (file_id,),
            )
            records = cursor.fetchall()

            cursor.close()
            connection.close()

            # Convertir les dates
            if file_info["upload_date"]:
                file_info["upload_date"] = file_info["upload_date"].isoformat()

            for record in records:
                if record["created_at"]:
                    record["created_at"] = record["created_at"].isoformat()

            return jsonify({"file_info": file_info, "records": records}), 200

        return jsonify({"error": "Erreur de connexion à la base de données"}), 500

    except Exception as e:
        return (
            jsonify(
                {
                    "error": "Erreur lors de la récupération des détails",
                    "details": str(e),
                }
            ),
            500,
        )


@app.route("/api/file/<file_id>", methods=["DELETE"])
def delete_file(file_id):
    """Supprimer un fichier et ses enregistrements"""
    try:
        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)

            # Récupérer les noms de fichiers
            cursor.execute(
                """
                SELECT original_filename, processed_filename
                FROM file_history WHERE file_id = %s
            """,
                (file_id,),
            )
            result = cursor.fetchone()

            if not result:
                return jsonify({"error": "Fichier non trouvé"}), 404

            # Supprimer de la base de données (cascade supprimera aussi processed_records)
            cursor.execute("DELETE FROM file_history WHERE file_id = %s", (file_id,))
            connection.commit()

            # Supprimer les fichiers physiques
            upload_path = os.path.join(
                app.config["UPLOAD_FOLDER"], f"{file_id}_{result['original_filename']}"
            )
            if os.path.exists(upload_path):
                os.remove(upload_path)

            if result["processed_filename"]:
                processed_path = os.path.join(
                    app.config["PROCESSED_FOLDER"], result["processed_filename"]
                )
                if os.path.exists(processed_path):
                    os.remove(processed_path)

            cursor.close()
            connection.close()

            return jsonify({"message": "Fichier supprimé avec succès"}), 200

        return jsonify({"error": "Erreur de connexion à la base de données"}), 500

    except Exception as e:
        return (
            jsonify({"error": "Erreur lors de la suppression", "details": str(e)}),
            500,
        )


# ==================== INITIALISATION ====================

if __name__ == "__main__":
    print("Initialisation de la base de données...")
    init_database()
    print("Démarrage du serveur Flask...")
    app.run(debug=True, host="0.0.0.0", port=5000)
